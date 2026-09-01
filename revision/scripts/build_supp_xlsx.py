#!/usr/bin/env python3
"""Write the supplemental CSVs as one workbook (Index + one sheet per table).
Numeric cells are stored as numbers (thousands separators stripped)."""
import csv, re, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

SUPP = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("supplementary")
OUT = SUPP / "TransGenic_supplementary_tables.xlsx"
ORDER = ["S1", "S2", "S3", "S4a", "S4b", "S4c", "S4d", "S5", "S6", "S7", "S8", "S9", "S10", "S11"]
NUM_INT = re.compile(r"^-?\d{1,3}(,\d{3})+$|^-?\d+$")
NUM_FLOAT = re.compile(r"^-?\d*\.\d+$")

def coerce(v):
    s = v.strip()
    if NUM_INT.match(s):
        return int(s.replace(",", ""))
    if NUM_FLOAT.match(s):
        return float(s)
    return v

md = (SUPP / "supplemental_information.md").read_text(encoding="utf-8")
titles = {m.group(1): m.group(2).strip() for m in re.finditer(r"^## Table (S\d+[a-d]?)\. (.+)$", md, re.M)}
wb = Workbook()
idx = wb.active; idx.title = "Index"
idx["A1"] = "TransGenic — supplemental tables"; idx["A1"].font = Font(bold=True)
for j, h in enumerate(["Table", "Title", "Data rows", "Source CSV"], 1):
    c = idx.cell(row=3, column=j, value=h); c.font = Font(bold=True)
for k, w in zip("ABCD", (12, 78, 11, 34)):
    idx.column_dimensions[k].width = w
idx.freeze_panes = "A4"
for i, sid in enumerate(ORDER):
    csvs = sorted(SUPP.glob(f"Table{sid}_*.csv"))
    assert len(csvs) == 1, (sid, csvs)
    rows = list(csv.reader(csvs[0].open(encoding="utf-8")))
    assert sid in titles, f"no title for Table {sid} in supplemental_information.md"
    ws = wb.create_sheet(f"Table {sid}")
    ws["A1"] = f"Table {sid}. {titles[sid]}"; ws["A1"].font = Font(bold=True)
    for j, h in enumerate(rows[0], 1):
        ws.cell(row=3, column=j, value=h).font = Font(bold=True)
    for r, row in enumerate(rows[1:], 4):
        for j, v in enumerate(row, 1):
            ws.cell(row=r, column=j, value=coerce(v))
    for j in range(1, len(rows[0]) + 1):
        width = max(len(str(row[j - 1])) if j - 1 < len(row) else 0 for row in rows)
        ws.column_dimensions[ws.cell(row=3, column=j).column_letter].width = max(9, min(60, width + 2))
    ws.freeze_panes = "A4"
    idx.append([]) if i == 0 and idx.max_row < 3 else None
    idx.cell(row=4 + i, column=1, value=f"Table {sid}")
    idx.cell(row=4 + i, column=2, value=titles[sid])
    idx.cell(row=4 + i, column=3, value=len(rows) - 1)
    idx.cell(row=4 + i, column=4, value=csvs[0].name)
wb.save(OUT)
print(f"wrote {OUT}: {len(wb.sheetnames)} sheets")
